# # -*- coding: utf-8 -*-
from openpyxl import load_workbook
import time
import help1


class shangChuan(object):
    def __init__(self, filename, filename1):
        # filename是土温汇总，filename1是上传模板
        self.filename = filename
        self.filename1 = filename1

    def solveFile(self):
        try:
            wb = load_workbook(self.filename)
            wc = load_workbook(self.filename1)
        except FileNotFoundError:
            return False
        sheetname = wb.sheetnames[0]  # 竖
        sheetname1 = wc.sheetnames[0]  # 6号连接通道测温
        curTime = time.strftime("%m/%d", time.localtime())
        rc = help1.getRowCol(self.filename, sheetname)
        rc1 = help1.getRowCol(self.filename1, sheetname1)
        # 从土温汇总中获取今日土温值
        data = help1.getColValue(self.filename, sheetname, rc[1] - 1)
        sheet1 = wc.get_sheet_by_name(sheetname1)
        total = rc1[0]
        count = 0
        r = 0
        # 将获取到的数据写入数据汇总中
        while (count < total):
            if data[r] != 42:
                sheet1.cell(row=count + 1, column=3).value = data[r]
                count = count + 1
            r = r + 1
        wc.save(self.filename1)
        return True
