# # -*- coding: utf-8 -*-
# 注意，使用前一定要将文件夹的名字改为今日时间，如10.29
from openpyxl import load_workbook
import time
import help1


class UpdateSoil(object):
    def __init__(self, filename):
        self.filename = filename

    def solveFile(self):
        try:
            wb = load_workbook(self.filename)
        except FileNotFoundError:
            return False
        sheetname1 = wb.sheetnames[0]
        sheetname2 = wb.sheetnames[1]
        sheetname3 = wb.sheetnames[2]
        # numrow代表今日土温总数127个数据，其中包括开头的空格
        # numcol代表经过excel处理的土温数据所在列号
        [numrow, numcol] = help1.getRowCol(self.filename, sheetname3)
        curTime = time.strftime("%m/%d", time.localtime())

        # 获取今日土温数据,注意,这里由于列的计数是从0开始,故要-1
        data = help1.getColValue(self.filename, sheetname3, numcol - 1)

        sheet1 = wb.get_sheet_by_name(sheetname1)
        sheet2 = wb.get_sheet_by_name(sheetname2)

        sheet1size = help1.getRowCol(self.filename, sheetname1)  # 获取1单元格行列数
        sheet2size = help1.getRowCol(self.filename, sheetname2)  # 获取2单元格行列数

        # 将sheet1、sheet2数据进行更新,使用openpyxl的时候，行数和列数都是从1开始计数，而data是从0开始计数
        sheet1.cell(row=1, column=sheet1size[1] + 1).value = curTime
        sheet2.cell(row=sheet2size[0] + 1, column=1).value = curTime
        updateCol = sheet1size[1] + 1
        updateRow = updateCol
        for r in range(numrow - 1):
            sheet1.cell(row=r + 2, column=updateCol).value = data[r + 1]
            sheet2.cell(row=updateRow, column=r + 2).value = data[r + 1]
            if data[r + 1] == 42:
                sheet1.cell(row=r + 2, column=updateCol).value = '#N/A'
                sheet2.cell(row=updateRow, column=r + 2).value = '#N/A'

        wb.save(self.filename)
        return True
